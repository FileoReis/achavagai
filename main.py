"""
main.py
AchaVagAI — encontra vagas de emprego combinando com o seu currículo.

Uso:
    python main.py caminho/para/curriculo.pdf

Fluxo:
  1. Lê o PDF do currículo. Se houver IA configurada (GEMINI_API_KEY ou
     ANTHROPIC_API_KEY, incluindo via arquivo .env), usa-a para extrair um perfil
     estruturado (cidade, bairro, cargo, senioridade, habilidades); caso contrário,
     usa uma extração local por palavras-chave.
  2. Pergunta/confirma os filtros de busca (cidade, bairro, palavra-chave — aceita
     várias separadas por vírgula, nº de dias).
  3. Busca vagas no LinkedIn, Indeed, Vagas.com, InfoJobs e nos portais regionais
     (RioVagas, Rio Emprega, Rio Empregos, VagasRio), uma vez para cada palavra-chave.
  4. Descarta vagas mais antigas que o número de dias informado (quando é possível
     identificar a data de publicação) e ranqueia as demais por aderência ao
     currículo (IA quando disponível, com justificativa por vaga; TF-IDF local como
     alternativa), descartando também as vagas abaixo da nota mínima configurada.
  5. Gera mensagens de candidatura personalizadas para as melhores vagas (se IA
     disponível).
  6. Salva os resultados em uma planilha Excel (.xlsx) formatada, com links
     clicáveis, e mostra as melhores vagas no terminal, em formato colorido.
"""

import argparse
import glob
import os
import re
import sys
from datetime import datetime

from dateutil import parser as dateparser

from dotenv import load_dotenv

load_dotenv()  # lê variáveis de um arquivo .env na pasta do projeto, se existir

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.prompt import Prompt, Confirm

from resume_parser import analisar_curriculo
from scrapers import (
    buscar_vagas_linkedin,
    buscar_vagas_indeed,
    buscar_vagas_todos_wordpress,
    buscar_vagas_vagas_com,
    buscar_vagas_infojobs,
)
from matcher import ranquear_por_similaridade, ranquear
from mensagem import gerar_mensagens_top
from config import SITES_ATIVOS, CIDADES_RJ, NOTA_MINIMA_PADRAO, UF_ALVO
import ia

console = Console()

SEPARADOR_TITULO_RE = re.compile(r"\s[–-]\s")


def dividir_palavras_chave(texto: str) -> list[str]:
    """Permite buscar vários cargos de uma vez, separados por vírgula."""
    termos = [t.strip() for t in texto.split(",")]
    return [t for t in termos if t]


def perguntar(pergunta: str, padrao: str | None) -> str:
    return Prompt.ask(f"[bold cyan]{pergunta}[/bold cyan]", default=padrao or "")


def _limpar_caminho(caminho: str) -> str:
    """Remove aspas que costumam vir junto quando se arrasta um arquivo para o
    terminal, ou quando se copia um caminho do Explorador de Arquivos do Windows."""
    return caminho.strip().strip('"').strip("'")


def resolver_caminho_pdf(caminho_informado: str | None) -> str:
    """Garante um caminho válido para o PDF do currículo antes de continuar —
    nunca deixa o script quebrar com um erro feio de arquivo não encontrado.

    - Se `caminho_informado` já existir, usa ele direto.
    - Se não existir (ou não tiver sido informado), tenta ajudar:
      encontra .pdf na pasta atual automaticamente (sugerindo se houver 1, ou
      listando para escolher se houver vários) e, em último caso, pede o
      caminho completo digitado."""
    if caminho_informado:
        caminho_informado = _limpar_caminho(caminho_informado)
        if os.path.isfile(caminho_informado):
            return caminho_informado
        console.print(f"[yellow]Não encontrei o arquivo \"{caminho_informado}\".[/yellow]")

    while True:
        pdfs_na_pasta = sorted(glob.glob("*.pdf"))

        if len(pdfs_na_pasta) == 1:
            if Confirm.ask(f"Encontrei [bold]{pdfs_na_pasta[0]}[/bold] na pasta atual — usar esse arquivo?", default=True):
                return pdfs_na_pasta[0]
        elif len(pdfs_na_pasta) > 1:
            console.print("\n[bold]Encontrei vários PDFs na pasta atual:[/bold]")
            for i, nome in enumerate(pdfs_na_pasta, start=1):
                console.print(f"  {i}. {nome}")
            escolha = _limpar_caminho(
                perguntar("Digite o número do currículo desejado, ou cole o caminho completo de outro arquivo", "1")
            )
            if escolha.isdigit() and 1 <= int(escolha) <= len(pdfs_na_pasta):
                return pdfs_na_pasta[int(escolha) - 1]
            if os.path.isfile(escolha):
                return escolha
            console.print(f"[yellow]Não encontrei \"{escolha}\".[/yellow]")
            continue

        caminho_manual = _limpar_caminho(perguntar("Digite o caminho completo do PDF do currículo", ""))
        if os.path.isfile(caminho_manual):
            return caminho_manual
        console.print(f"[red]Não encontrei \"{caminho_manual}\". Tente novamente.[/red]")


def _salvar_env(chaves: dict) -> None:
    """Salva (ou atualiza) o arquivo .env com as chaves informadas, preservando
    outras variáveis que já estivessem lá."""
    caminho_env = ".env"
    linhas_existentes = {}
    if os.path.isfile(caminho_env):
        with open(caminho_env, "r", encoding="utf-8") as f:
            for linha in f:
                if "=" in linha and not linha.strip().startswith("#"):
                    chave, _, valor = linha.strip().partition("=")
                    linhas_existentes[chave] = valor

    linhas_existentes.update(chaves)

    with open(caminho_env, "w", encoding="utf-8") as f:
        for chave, valor in linhas_existentes.items():
            f.write(f"{chave}={valor}\n")

    console.print(f"[bold green]✓[/bold green] Chave(s) salva(s) em [bold].env[/bold] — não vai precisar configurar de novo.")


def configurar_ia_interativa() -> None:
    """Se nenhuma chave de IA estiver configurada (nem por .env, nem por
    variável de ambiente), explica os benefícios e oferece configurar uma
    agora — sem precisar sair do script para editar arquivo nenhum."""
    if ia.provedor_disponivel():
        return

    console.print(Panel(
        "O AchaVagAI funciona sem IA (usando um ranking local mais simples, por "
        "similaridade de texto), mas fica bem melhor com uma IA configurada:\n\n"
        "• Lê o currículo inteiro e entende contexto, não só palavras-chave soltas\n"
        "• Avalia requisitos reais da vaga (ex.: categoria de CNH, curso técnico, "
        "vaga exclusiva PCD) em vez de só comparar se o título parece parecido\n"
        "• Explica o motivo de cada nota\n"
        "• Pode gerar mensagens de candidatura prontas para copiar e enviar\n\n"
        "O [bold]Gemini[/bold] (Google) é gratuito e não pede cartão de crédito — "
        "crie uma chave em [bold]https://aistudio.google.com/apikey[/bold].",
        title="[bold]Nenhuma IA configurada[/bold]",
        border_style="yellow",
    ))

    if not Confirm.ask("Quer configurar uma chave de IA agora?", default=True):
        console.print("[dim]Ok, seguindo sem IA. Pode configurar depois criando um arquivo .env na pasta do projeto.[/dim]")
        return

    nomes_provedor = {"gemini": ("GEMINI_API_KEY", "Gemini"), "claude": ("ANTHROPIC_API_KEY", "Claude")}
    chaves_coletadas = {}

    while True:
        escolha = Prompt.ask(
            "Qual provedor você quer configurar",
            choices=["gemini", "claude"],
            default="gemini",
        )
        nome_var, nome_exibicao = nomes_provedor[escolha]
        chave_valor = Prompt.ask(f"Cole sua chave do {nome_exibicao}").strip()

        if chave_valor:
            chaves_coletadas[nome_var] = chave_valor
            os.environ[nome_var] = chave_valor
            console.print(f"[green]✓[/green] {nome_exibicao} configurado para esta execução.")
        else:
            console.print("[yellow]Nenhuma chave informada — pulando.[/yellow]")

        if not Confirm.ask("Quer adicionar outra chave também (ex.: como alternativa)?", default=False):
            break

    if chaves_coletadas:
        _salvar_env(chaves_coletadas)


def montar_filtros(perfil) -> dict:
    console.print("\n[bold]--- Confirme ou ajuste os filtros de busca ---[/bold] [dim](Enter mantém o valor sugerido)[/dim]")
    cidade = perguntar("Cidade", perfil.cidade)
    bairro = perguntar("Bairro (opcional, usado apenas na exibição)", perfil.bairro)
    palavra_chave = perguntar(
        "Palavra-chave / cargo desejado [dim](pode informar vários, separados por vírgula)[/dim]",
        perfil.cargo_sugerido,
    )
    dias = perguntar("Considerar vagas publicadas até quantos dias atrás?", "30")
    return {"cidade": cidade, "bairro": bairro, "palavra_chave": palavra_chave, "dias": dias}


def buscar_todas_vagas(filtros: dict) -> list:
    termos = dividir_palavras_chave(filtros["palavra_chave"]) or [""]
    todas_vagas = []

    fontes = [
        ("linkedin", "LinkedIn", buscar_vagas_linkedin),
        ("indeed", "Indeed", buscar_vagas_indeed),
        ("vagas_com", "Vagas.com", buscar_vagas_vagas_com),
        ("infojobs", "InfoJobs", buscar_vagas_infojobs),
    ]

    for termo in termos:
        rotulo = termo or "(sem palavra-chave)"

        for chave_config, nome_site, funcao_busca in fontes:
            if not SITES_ATIVOS.get(chave_config):
                continue
            with console.status(f"[cyan]Buscando no {nome_site} — {rotulo}...[/cyan]"):
                vagas = funcao_busca(termo, filtros["cidade"])
            for vaga in vagas:
                vaga.termo_busca = termo
            if vagas:
                console.print(f"  [green]✓[/green] {nome_site} ({rotulo}): [bold]{len(vagas)}[/bold] vaga(s)")
            todas_vagas += vagas

        with console.status(f"[cyan]Buscando nos portais regionais — {rotulo}...[/cyan]"):
            vagas = buscar_vagas_todos_wordpress(termo, SITES_ATIVOS)
        for vaga in vagas:
            vaga.termo_busca = termo
        console.print(f"  [green]✓[/green] Portais regionais ({rotulo}): [bold]{len(vagas)}[/bold] vaga(s)")
        todas_vagas += vagas

    return todas_vagas


def remover_duplicadas(vagas: list) -> list:
    vistos = set()
    unicas = []
    for vaga in vagas:
        if vaga.link not in vistos:
            vistos.add(vaga.link)
            unicas.append(vaga)
    return unicas


UF_NA_CIDADE_RE = re.compile(r"/\s*([A-Z]{2})\b")


def filtrar_por_regiao(vagas: list, uf_alvo: str = UF_ALVO) -> tuple[list, int]:
    """Descarta vagas cujo campo `cidade` traga explicitamente um estado (UF)
    diferente do configurado — ex.: "São Paulo / SP" é descartado se `uf_alvo` for
    "RJ". Isso é feito no código (não só confiando na IA) porque a análise de IA
    tratava localização apenas como uma preferência, não uma exclusão — e vagas de
    outros estados chegavam a passar com nota razoável mesmo assim.

    Vagas sem UF identificável no campo cidade (a maioria, já que muitos sites não
    estruturam isso) NÃO são descartadas por este filtro — segue tudo para a IA/
    TF-IDF avaliar normalmente. Retorna (vagas_filtradas, quantidade_removida)."""
    filtradas = []
    removidas = 0
    for vaga in vagas:
        match = UF_NA_CIDADE_RE.search(vaga.cidade or "")
        if match and match.group(1).upper() != uf_alvo.upper():
            removidas += 1
            continue
        filtradas.append(vaga)
    return filtradas, removidas


def _idade_em_dias(data_publicacao: str) -> int | None:
    """Calcula há quantos dias a vaga foi publicada. Retorna None se a data
    estiver vazia ou em um formato não reconhecido (nesse caso, a vaga não é
    descartada pelo filtro de data — não dá pra confirmar que está velha)."""
    if not data_publicacao:
        return None
    try:
        data = dateparser.parse(data_publicacao)
    except (ValueError, OverflowError, TypeError):
        return None
    if data is None:
        return None
    agora = datetime.now(data.tzinfo) if data.tzinfo else datetime.now()
    return (agora - data).days


def filtrar_por_data(vagas: list, dias_maximo: int) -> tuple[list, int, int]:
    """Remove vagas cuja data de publicação seja mais antiga que `dias_maximo`.
    Vagas sem data identificável são mantidas (não há como confirmar a idade),
    mas contadas separadamente para informar o usuário.
    Retorna (vagas_filtradas, quantidade_removida_por_idade, quantidade_sem_data)."""
    filtradas = []
    removidas_por_idade = 0
    sem_data = 0

    for vaga in vagas:
        idade = _idade_em_dias(vaga.data_publicacao)
        if idade is None:
            sem_data += 1
            filtradas.append(vaga)
        elif idade <= dias_maximo:
            filtradas.append(vaga)
        else:
            removidas_por_idade += 1

    return filtradas, removidas_por_idade, sem_data


def enriquecer_vaga(vaga):
    """Tenta preencher empresa/cidade quando vieram vazias, usando o padrão
    "Cargo – Empresa – Local" do próprio título e a lista de cidades conhecidas."""
    partes = SEPARADOR_TITULO_RE.split(vaga.titulo)

    if not vaga.empresa and len(partes) >= 3:
        vaga.empresa = partes[1].strip()

    if not vaga.cidade:
        candidata = partes[-1].strip() if len(partes) >= 2 else ""
        if "RJ" in candidata or any(cidade.lower() in candidata.lower() for cidade in CIDADES_RJ):
            vaga.cidade = candidata
        else:
            cidade_no_titulo = next(
                (cidade for cidade in CIDADES_RJ if cidade.lower() in vaga.titulo.lower()), None
            )
            if cidade_no_titulo:
                vaga.cidade = cidade_no_titulo

    return vaga


def salvar_excel(ranking: list, mensagens: list, caminho: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vagas"

    colunas = ["Pontuação (%)", "Cargo / Título", "Empresa", "Cidade", "Publicada em", "Fonte", "Motivo (IA)", "Link"]
    ws.append(colunas)
    _formatar_cabecalho(ws, len(colunas))

    for vaga, pontuacao, motivo in ranking:
        vaga = enriquecer_vaga(vaga)
        ws.append([pontuacao, vaga.titulo, vaga.empresa, vaga.cidade, vaga.data_publicacao, vaga.fonte, motivo, vaga.link])
        numero_linha = ws.max_row
        celula_link = ws.cell(row=numero_linha, column=8)
        celula_link.hyperlink = vaga.link
        celula_link.font = Font(color="0563C1", underline="single")
        ws.cell(row=numero_linha, column=1).alignment = Alignment(horizontal="center")

    larguras = {1: 14, 2: 50, 3: 26, 4: 20, 5: 16, 6: 14, 7: 45, 8: 55}
    for col_idx, largura in larguras.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = largura
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    if mensagens:
        ws2 = wb.create_sheet("Mensagens de Candidatura")
        colunas2 = ["Cargo / Título", "Link", "Mensagem sugerida"]
        ws2.append(colunas2)
        _formatar_cabecalho(ws2, len(colunas2))
        for vaga, _pontuacao, mensagem in mensagens:
            ws2.append([vaga.titulo, vaga.link, mensagem])
            numero_linha = ws2.max_row
            celula_link = ws2.cell(row=numero_linha, column=2)
            celula_link.hyperlink = vaga.link
            celula_link.font = Font(color="0563C1", underline="single")
            ws2.cell(row=numero_linha, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.column_dimensions["A"].width = 45
        ws2.column_dimensions["B"].width = 50
        ws2.column_dimensions["C"].width = 90
        ws2.freeze_panes = "A2"

    wb.save(caminho)


def _formatar_cabecalho(ws, n_colunas: int) -> None:
    cabecalho_fonte = Font(bold=True, color="FFFFFF")
    cabecalho_fundo = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    for col_idx in range(1, n_colunas + 1):
        celula = ws.cell(row=1, column=col_idx)
        celula.font = cabecalho_fonte
        celula.fill = cabecalho_fundo
        celula.alignment = Alignment(horizontal="center", vertical="center")


def exibir_perfil(perfil) -> None:
    tabela = Table(show_header=False, box=None, padding=(0, 1))
    tabela.add_row("[bold]Nome:[/bold]", perfil.nome or "[dim]não identificado[/dim]")
    tabela.add_row("[bold]Cidade:[/bold]", perfil.cidade or "[dim]não identificada[/dim]")
    tabela.add_row("[bold]Bairro:[/bold]", perfil.bairro or "[dim]não identificado[/dim]")
    tabela.add_row("[bold]Habilidades:[/bold]", ", ".join(perfil.habilidades) or "[dim]nenhuma identificada[/dim]")
    tabela.add_row("[bold]Cargo sugerido:[/bold]", perfil.cargo_sugerido or "[dim]não identificado[/dim]")
    if perfil.senioridade:
        tabela.add_row("[bold]Senioridade:[/bold]", perfil.senioridade)
    if perfil.resumo:
        tabela.add_row("[bold]Resumo:[/bold]", perfil.resumo)
    titulo = "Currículo analisado" + (" (via IA)" if perfil.extraido_por_ia else " (extração local)")
    console.print(Panel(tabela, title=f"[bold white]{titulo}[/bold white]", border_style="cyan"))
    if perfil.erro_ia:
        console.print(f"[yellow]A extração por IA do currículo falhou — usei a extração local acima. Motivo: {perfil.erro_ia}[/yellow]")


def exibir_resultados(ranking: list, top_n: int) -> None:
    tabela = Table(title=f"Top {min(top_n, len(ranking))} vagas mais compatíveis", header_style="bold white on dark_cyan")
    tabela.add_column("Nota", justify="center", width=6)
    tabela.add_column("Vaga", style="bold")
    tabela.add_column("Local", style="magenta")
    tabela.add_column("Fonte", style="yellow")
    tabela.add_column("Motivo (IA)", style="dim italic")

    for vaga, pontuacao, motivo in ranking[:top_n]:
        vaga = enriquecer_vaga(vaga)
        cor_nota = "green" if pontuacao >= 60 else ("yellow" if pontuacao >= 35 else "red")
        tabela.add_row(f"[{cor_nota}]{pontuacao:.0f}[/{cor_nota}]", vaga.titulo, vaga.cidade or "-", vaga.fonte, motivo or "-")
    console.print(tabela)

    console.print("\n[bold]Links para candidatura:[/bold]")
    for i, (vaga, _pontuacao, _motivo) in enumerate(ranking[:top_n], start=1):
        console.print(f"  {i}. [link={vaga.link}]{vaga.link}[/link]")


def exibir_mensagens(mensagens: list) -> None:
    if not mensagens:
        return
    console.print("\n[bold magenta]--- Mensagens de candidatura sugeridas (top vagas) ---[/bold magenta]")
    for vaga, _pontuacao, mensagem in mensagens:
        console.print(Panel(mensagem, title=f"[bold]{vaga.titulo}[/bold]", border_style="magenta", subtitle=vaga.link))


def main():
    parser = argparse.ArgumentParser(description="AchaVagAI - encontra vagas combinando com seu currículo")
    parser.add_argument("curriculo_pdf", nargs="?", default=None, help="Caminho para o arquivo PDF do currículo (se não informado, o script ajuda a encontrar)")
    parser.add_argument("--sem-ia", action="store_true", help="Desativa toda a análise de IA (ranking, extração de perfil e mensagens), mesmo com chave configurada.")
    parser.add_argument("--com-mensagens", action="store_true", help="Gera mensagens de candidatura personalizadas para as melhores vagas (desativado por padrão — usa chamadas extras de IA).")
    parser.add_argument("--top", type=int, default=15, help="Quantas vagas mostrar no resumo final (padrão: 15)")
    parser.add_argument("--top-mensagens", type=int, default=5, help="Para quantas vagas do topo gerar mensagem de candidatura (padrão: 5)")
    parser.add_argument("--nota-minima", type=float, default=NOTA_MINIMA_PADRAO, help=f"Nota mínima (0-100) para uma vaga aparecer nos resultados (padrão: {NOTA_MINIMA_PADRAO})")
    args = parser.parse_args()

    if args.sem_ia:
        os.environ.pop("GEMINI_API_KEY", None)
        os.environ.pop("ANTHROPIC_API_KEY", None)

    console.rule("[bold cyan]ACHAVAGAI[/bold cyan]")

    if not args.sem_ia:
        configurar_ia_interativa()

    caminho_pdf = resolver_caminho_pdf(args.curriculo_pdf)

    console.print(f"[dim]Lendo currículo:[/dim] {caminho_pdf}")
    perfil = analisar_curriculo(caminho_pdf)
    exibir_perfil(perfil)

    filtros = montar_filtros(perfil)

    console.print("\n[bold]--- Buscando vagas ---[/bold]")
    vagas = remover_duplicadas(buscar_todas_vagas(filtros))
    console.print(f"\n[bold]Total de vagas encontradas (após remover duplicadas):[/bold] {len(vagas)}")

    if not vagas:
        console.print("[red]Nenhuma vaga encontrada. Tente ajustar a palavra-chave ou verifique sua conexão.[/red]")
        sys.exit(0)

    try:
        dias_maximo = int(filtros["dias"])
    except (ValueError, TypeError):
        dias_maximo = None

    if dias_maximo:
        vagas, removidas_por_idade, sem_data = filtrar_por_data(vagas, dias_maximo)
        if removidas_por_idade:
            console.print(
                f"[dim]{removidas_por_idade} vaga(s) publicada(s) há mais de {dias_maximo} dias foram descartadas.[/dim]"
            )
        if sem_data:
            console.print(
                f"[dim]{sem_data} vaga(s) sem data identificável foram mantidas (não dá pra confirmar a idade).[/dim]"
            )

    if not vagas:
        console.print("[red]Nenhuma vaga passou do filtro de data. Tente aumentar o número de dias.[/red]")
        sys.exit(0)

    vagas, removidas_por_regiao = filtrar_por_regiao(vagas)
    if removidas_por_regiao:
        console.print(
            f"[dim]{removidas_por_regiao} vaga(s) de fora do estado {UF_ALVO} foram descartadas "
            f"(ajustável em UF_ALVO no config.py).[/dim]"
        )

    if not vagas:
        console.print("[red]Nenhuma vaga passou do filtro de região. Tente ajustar UF_ALVO no config.py.[/red]")
        sys.exit(0)

    provedor = ia.provedor_disponivel()
    usou_ia = False
    if provedor:
        nome_provedor = "Gemini" if provedor == "gemini" else "Claude"
        console.print(f"\n[bold magenta]Analisando vagas com IA ({nome_provedor}) — isso pode levar alguns segundos...[/bold magenta]")
        with console.status("[magenta]Avaliando aderência ao currículo...[/magenta]"):
            ranking, usou_ia = ranquear(perfil.texto_completo, vagas)
        if not usou_ia:
            motivo_erro = ia.ultimo_erro or "motivo desconhecido"
            console.print(
                f"[yellow]A chamada à IA ({nome_provedor}) falhou nesta execução — usando ranking local "
                f"(TF-IDF) como alternativa, sem justificativa por vaga.[/yellow]\n"
                f"[dim]Motivo reportado: {motivo_erro}[/dim]"
            )
    else:
        console.print(
            "\n[yellow]Nenhuma chave de IA configurada — usando ranking local (TF-IDF), mais simples.[/yellow]\n"
            "[dim]Para uma análise mais profunda e gratuita, crie uma chave em "
            "https://aistudio.google.com/apikey, salve em um arquivo .env como GEMINI_API_KEY=sua-chave.[/dim]"
        )
        ranking = ranquear_por_similaridade(perfil.texto_completo, vagas)

    total_antes_filtro = len(ranking)
    ranking_completo = ranking  # guarda a lista completa (antes do filtro) para diagnóstico, se precisar
    ranking = [item for item in ranking if item[1] >= args.nota_minima]
    descartadas = total_antes_filtro - len(ranking)
    if descartadas:
        console.print(f"[dim]{descartadas} vaga(s) abaixo de {args.nota_minima:.0f}% de aderência foram descartadas (ajuste com --nota-minima).[/dim]")

    if not ranking:
        console.print(
            f"[red]Nenhuma vaga passou do filtro de nota mínima ({args.nota_minima:.0f}%).[/red]\n"
            "[yellow]Isso pode ser normal — ex.: se a vaga exige um requisito que seu currículo não "
            "atende (CNH de categoria diferente, curso técnico específico etc.), a IA reprova mesmo que "
            "o título pareça relacionado.[/yellow]\n"
            "[dim]Veja abaixo as vagas com maior nota mesmo assim, para entender o motivo:[/dim]"
        )
        exibir_resultados(ranking_completo, min(5, len(ranking_completo)))
        console.print(
            "\n[dim]Se quiser ver todas mesmo abaixo da nota mínima, rode de novo com --nota-minima 0.[/dim]"
        )
        nome_arquivo = f"vagas_encontradas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        salvar_excel(ranking_completo, [], nome_arquivo)
        console.print(f"\n[bold green]✓[/bold green] Mesmo assim, salvei todas as vagas avaliadas (com nota e motivo) em: [bold]{nome_arquivo}[/bold]")
        sys.exit(0)

    console.print()
    exibir_resultados(ranking, args.top)

    mensagens = []
    if provedor and args.com_mensagens:
        console.print(f"\n[bold magenta]Gerando mensagens de candidatura para as top {args.top_mensagens} vagas...[/bold magenta]")
        with console.status("[magenta]Escrevendo mensagens personalizadas...[/magenta]"):
            mensagens = gerar_mensagens_top(perfil.texto_completo, ranking, quantidade=args.top_mensagens)
        exibir_mensagens(mensagens)

    nome_arquivo = f"vagas_encontradas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    salvar_excel(ranking, mensagens, nome_arquivo)
    console.print(f"\n[bold green]✓[/bold green] Resultados completos salvos em: [bold]{nome_arquivo}[/bold]")
    console.print("[dim]Abra no Excel — os links já são clicáveis; mensagens de candidatura ficam na segunda aba.[/dim]")


if __name__ == "__main__":
    main()
